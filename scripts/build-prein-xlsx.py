#!/usr/bin/env python3
"""Generate one PreIn xlsx per SNKR shipment, configured by prein.config.json."""
import csv
import datetime as dt
import glob
import json
import os
import re
import sys
from copy import copy
from pathlib import Path

import openpyxl

CWD = Path.cwd()
CONFIG_PATH = CWD / "prein.config.json"
if not CONFIG_PATH.exists():
    sys.exit("prein.config.json missing. Copy prein.config.example.json and fill it in.")
CFG = json.loads(CONFIG_PATH.read_text())

def expand(p: str) -> Path:
    return Path(os.path.expanduser(p)).resolve()

TEMPLATE = expand(CFG["template"])
OUT_DIR = expand(CFG["out_dir"])
TEMPLATE_SHEET = CFG.get("template_sheet", "PRE_IN")
OUT_PREFIX = CFG.get("out_prefix", "PreIn_")
OUT_SUFFIX = CFG.get("out_suffix", "")
COLS = CFG["columns"]

SHIPMENTS = json.loads((CWD / "prein-shipments.json").read_text())

# ---------- Lookups ----------
cart_by_tx: dict[str, dict] = {}
for p in glob.glob(CFG.get("orders_csv_glob", "./orders-cart-*.csv")):
    try:
        with open(p) as f:
            for row in csv.DictReader(f):
                tx = (row.get("transaction_id") or "").strip()
                if not tx or tx in cart_by_tx:
                    continue
                cart_by_tx[tx] = {
                    "cert": (row.get("psa_cert") or "").strip(),
                    "grade_str": (row.get("grade") or "").strip(),
                    "name": (row.get("name") or "").strip(),
                }
    except Exception as e:
        print(f"  ⚠ couldn't read {p}: {e}", file=sys.stderr)

psa_by_tx: dict[str, dict] = {}
GRADE_NAME_TO_NUM = {
    "GEM MT 10": 10, "GEM MINT 10": 10, "MINT 9": 9, "NM-MT 8": 8, "NM-MT+ 8.5": 8.5,
    "NM 7": 7, "EX-MT 6": 6, "EX 5": 5, "VG-EX 4": 4, "VG 3": 3, "GOOD 2": 2, "PR 1": 1,
    "AUTHENTIC": 0,
}
for p in glob.glob(CFG.get("psa_verify_glob", "./psa-verify-*.json")):
    try:
        for rec in json.loads(Path(p).read_text()):
            tx = rec.get("tx")
            txt = rec.get("psaText") or ""
            if not tx:
                continue
            grade_num = None
            m = re.search(r"ITEM GRADE\s*\|\s*([A-Z+\- ]+\d+(?:\.\d+)?)", txt)
            if m:
                grade_num = GRADE_NAME_TO_NUM.get(m.group(1).strip())
                if grade_num is None:
                    gm = re.search(r"(\d+(?:\.\d+)?)", m.group(1))
                    if gm: grade_num = float(gm.group(1))
            desc = None
            cert = rec.get("cert") or ""
            if cert:
                dm = re.search(rf"#{cert}\s*\|\s*([^|]+?)\s*\|\s*ITEM GRADE", txt)
                if dm:
                    desc = dm.group(1).strip()
                    tail = re.search(r"#\d+\s+(.*)", desc)
                    if tail: desc = tail.group(1).strip()
            psa_by_tx[tx] = {"grade_num": grade_num, "desc": desc}
    except Exception as e:
        print(f"  ⚠ couldn't read {p}: {e}", file=sys.stderr)

def short_name(full: str) -> str:
    if not full: return ""
    return re.split(r"[\[\(]", full)[0].strip().upper()

def grade_str_to_num(s: str):
    if not s: return None
    m = re.search(r"PSA\s*(\d+(?:\.\d+)?)", s)
    return float(m.group(1)) if m else None

def parse_date(iso):
    if not iso: return dt.datetime.now()
    try:
        return dt.datetime.fromisoformat(iso.replace("Z", "+00:00")).replace(tzinfo=None)
    except Exception:
        return dt.datetime.now()

# ---------- Generate ----------
if not TEMPLATE.exists():
    sys.exit(f"template missing: {TEMPLATE}")
OUT_DIR.mkdir(parents=True, exist_ok=True)

written, skipped, missing = 0, 0, 0
for tracking, ship in SHIPMENTS.items():
    out_path = OUT_DIR / f"{OUT_PREFIX}{tracking}{OUT_SUFFIX}.xlsx"
    if out_path.exists():
        skipped += 1
        continue

    wb = openpyxl.load_workbook(TEMPLATE)
    if TEMPLATE_SHEET not in wb.sheetnames:
        sys.exit(f"sheet {TEMPLATE_SHEET!r} not found in template. Available: {wb.sheetnames}")
    ws = wb[TEMPLATE_SHEET]

    styles = []
    for col in range(1, ws.max_column + 1):
        c = ws.cell(row=2, column=col)
        styles.append({
            "font": copy(c.font), "fill": copy(c.fill), "border": copy(c.border),
            "alignment": copy(c.alignment), "number_format": c.number_format,
        })
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    date_obj = parse_date(ship.get("date"))
    rows_written = 0
    for r in ship["rows"]:
        tx = r.get("tx") or ""
        cart = cart_by_tx.get(tx, {})
        psa = psa_by_tx.get(tx, {})
        cert = r.get("cert") or cart.get("cert")
        if not cert:
            print(f"  ⚠ no cert for {tracking}/{r.get('orderId')} tx={tx}", file=sys.stderr)
            missing += 1
            continue
        grade_num = psa.get("grade_num") or grade_str_to_num(cart.get("grade_str", "") or r.get("grade", ""))
        eng_name = psa.get("desc") or short_name(cart.get("name") or r.get("name") or "")

        i = rows_written + 2
        values = {}
        if "customer" in COLS:      values[COLS["customer"]] = CFG.get("customer", "")
        if "inbound_date" in COLS:  values[COLS["inbound_date"]] = date_obj
        if "tracking" in COLS:      values[COLS["tracking"]] = tracking
        if "warehouse" in COLS:     values[COLS["warehouse"]] = CFG.get("warehouse", "")
        if "remarks" in COLS:       values[COLS["remarks"]] = f"PSA Cert: {cert}"
        if "item_no" in COLS:       values[COLS["item_no"]] = f"PSA-{cert}"
        if "collection" in COLS:    values[COLS["collection"]] = CFG.get("collection", "")
        if "card_name" in COLS:     values[COLS["card_name"]] = eng_name
        if "grader" in COLS:        values[COLS["grader"]] = CFG.get("grader", "PSA")
        if "grade" in COLS:         values[COLS["grade"]] = grade_num
        if "pack_size" in COLS:     values[COLS["pack_size"]] = 1
        if "qty_pcs" in COLS:       values[COLS["qty_pcs"]] = 1
        if "qty_ctns" in COLS:      values[COLS["qty_ctns"]] = 1
        if "qty_pcs_unit" in COLS:  values[COLS["qty_pcs_unit"]] = "pcs"
        if "qty_ctns_unit" in COLS: values[COLS["qty_ctns_unit"]] = "CTN"
        if "id" in COLS:            values[COLS["id"]] = int(cert) if str(cert).isdigit() else cert
        if "serial" in COLS:        values[COLS["serial"]] = f"PSA-{cert}"
        if "redeemed" in COLS:      values[COLS["redeemed"]] = "no"

        for col in range(1, len(styles) + 1):
            cell = ws.cell(row=i, column=col, value=values.get(col))
            s = styles[col - 1]
            cell.font = copy(s["font"]); cell.fill = copy(s["fill"])
            cell.border = copy(s["border"]); cell.alignment = copy(s["alignment"])
            cell.number_format = s["number_format"]
        rows_written += 1

    if rows_written == 0:
        print(f"  ⚠ {tracking}: no usable rows, not writing file", file=sys.stderr)
        continue

    wb.save(out_path)
    written += 1
    print(f"✓ {out_path.name}  ({rows_written} items)")

print(f"\n{written} new, {skipped} existing, {missing} rows skipped (missing cert)")
